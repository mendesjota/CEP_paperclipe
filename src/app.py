"""
CEP Paperclipe - Sistema de Validação e Correção de CEPs
Seguindo rigorosamente o idea.MD

4 Planilhas de Saída:
1. 1_ceps_validos.xlsx - Dados originais no modelo Correios
2. 2_ceps_corrigidos.xlsx - Dados atualizados no modelo Correios  
3. 3_ceps_nao_encontrados.xlsx - Log de erros
4. 4_base_consolidada.xlsx - Dados originais + modificado
"""
import io
import time
import threading
import urllib.parse
import concurrent.futures
import pandas as pd
import streamlit as st
import requests
import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="CEP Paperclipe", page_icon="📮")
st.title("📮 CEP Paperclipe")
st.markdown("### Validar, Corrigir CEPs e Gerar Lista para Envio")

# Colunas esperadas
CABEÇALHO_ESPERADO = ['CPF', 'MATRICULA', 'NOME', 'DATA NASCIMENTO', 'ENDERECO', 
                       'COMPLEMENTO DO ENDERECO', 'BAIRRO', 'MUNICIPIO', 'UF', 'CEP']

# Session state
if 'processado' not in st.session_state:
    st.session_state.processado = False
if 'resultados' not in st.session_state:
    st.session_state.resultados = None


# ============================================
# FUNÇÕES
# ============================================

def excel_to_bytes_text(df):
    """Gera Excel com todas as colunas em formato TEXTO (@)"""
    from openpyxl.styles import numbers
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='left')
            cell.number_format = numbers.FORMAT_TEXT
            if r_idx == 1:
                cell.font = Font(bold=True)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def excel_correios_format(rows):
    """Gera Excel no modelo Correios com formato personalizado (zeros)"""
    from openpyxl.styles import numbers
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Correios"
    
    # Cabeçalho
    headers = ['PREFIXO', 'NOME', 'CPF', 'CEP', 'LOGRADOURO', 'ENDERECO', 'BAIRRO', 'MUNICIPIO', 'UF', 'FLAG', 'COMPLEMENTO', 'NUMERO']
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.font = Font(bold=True)
    
    # Dados
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='left')
            
            # Formato especial para CPF (11 zeros) e CEP (8 zeros)
            if c_idx == 3:  # CPF
                cell.number_format = '00000000000'
            elif c_idx == 4:  # CEP
                cell.number_format = '00000000'
    
    # Ajustar largura
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 5
    ws.column_dimensions['K'].width = 40
    ws.column_dimensions['L'].width = 8
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def limpar_cep(cep):
    """Remove caracteres não numéricos"""
    return re.sub(r'\D', '', str(cep).strip())

def limpar_digitos(valor):
    """Remove tudo que não for dígito"""
    return re.sub(r'\D', '', str(valor).strip())

def normalizar_municipio(municipio):
    """Uppercase e espaços normalizados"""
    return re.sub(r'\s+', ' ', str(municipio or '').upper().strip())

def normalizar_cep(cep):
    """Normaliza CEP para 8 dígitos.
    - Completa com zeros à esquerda quando tem MENOS de 8 (perda de zero é comum).
    - NÃO trunca: se tiver MAIS de 8 dígitos, devolve como está para ser
      detectado como inválido (truncar cegamente geraria CEP de outro lugar).
    """
    cep_limpo = limpar_cep(cep)
    if len(cep_limpo) < 8:
        cep_limpo = cep_limpo.zfill(8)
    return cep_limpo

def cep_tem_formato_valido(cep):
    """CEP válido em formato = exatamente 8 dígitos numéricos."""
    return len(limpar_cep(cep)) == 8

def normalizar_cpf(cpf):
    """Completa CPF com zeros à esquerda até 11 dígitos (sem truncar)."""
    cpf_limpo = limpar_digitos(cpf)
    if len(cpf_limpo) < 11:
        cpf_limpo = cpf_limpo.zfill(11)
    return cpf_limpo

def cpf_valido(cpf):
    """Tamanho aceito pelo Correios: 11 (CPF) ou 14 (CNPJ)."""
    return len(limpar_digitos(cpf)) in (11, 14)

# Cache em memória das consultas ao DNE (evita repetir o mesmo CEP no lote)
_cache_dne = {}

# Sessão HTTP reutilizável (connection pooling) + identificação
_sessao = requests.Session()
_sessao.headers.update({"User-Agent": "CEP-Paperclipe/6.2 (uso interno)"})

# ---- Ritmo controlado (rate limit global) para NÃO tomar bloqueio das APIs ----
MIN_INTERVALO = 0.15            # segundos entre requisições (~6-7 req/s combinado)
COOLDOWN_S = 60                 # tempo de "descanso" de um provedor após bloqueio
FALHAS_P_COOLDOWN = 3           # falhas seguidas que colocam o provedor em descanso
_rate_lock = threading.Lock()
_rate_prox = {"t": 0.0}

def _aguardar_vez():
    """Espaça o INÍCIO das requisições entre todas as threads (gentil com a API)."""
    with _rate_lock:
        agora = time.time()
        espera = _rate_prox["t"] - agora
        if espera > 0:
            time.sleep(espera)
            agora = time.time()
        _rate_prox["t"] = max(agora, _rate_prox["t"]) + MIN_INTERVALO

# Estado por provedor: descanso (cooldown) ao tomar 429/timeout repetido
_provedores = {
    "viacep":    {"falhas": 0, "ate": 0.0},
    "brasilapi": {"falhas": 0, "ate": 0.0},
}
_rr = {"i": 0}  # revezamento de quem tenta primeiro

def _provedor_disponivel(nome):
    return time.time() >= _provedores[nome]["ate"]

def _descansar(nome, segundos=COOLDOWN_S):
    _provedores[nome]["ate"] = time.time() + segundos
    _provedores[nome]["falhas"] = 0

def _falha(nome):
    _provedores[nome]["falhas"] += 1
    if _provedores[nome]["falhas"] >= FALHAS_P_COOLDOWN:
        _descansar(nome)

def _get_provedor(nome, url, timeout=4):
    """GET respeitando o descanso do provedor e o ritmo global.
    Retorna o JSON (dict/list) em 200; None caso contrário."""
    if not _provedor_disponivel(nome):
        return None
    _aguardar_vez()
    try:
        r = _sessao.get(url, timeout=timeout)
        if r.status_code == 200:
            _provedores[nome]["falhas"] = 0
            return r.json()
        if r.status_code == 404:
            _provedores[nome]["falhas"] = 0  # respondeu: CEP não existe (não é falha)
            return None
        if r.status_code == 429:
            _descansar(nome)                 # bloqueio explícito -> descansa já
            return None
        _falha(nome)
        return None
    except Exception:
        _falha(nome)
        return None

def _consultar_viacep(cep):
    d = _get_provedor("viacep", f"https://viacep.com.br/ws/{cep}/json/")
    if isinstance(d, dict) and not d.get('erro'):
        return {
            "encontrado": True,
            "uf": (d.get('uf') or '').upper().strip(),
            "localidade": normalizar_municipio(d.get('localidade')),
            "logradouro": (d.get('logradouro') or '').upper().strip(),
            "bairro": (d.get('bairro') or '').upper().strip(),
        }
    return None

def _consultar_brasilapi(cep):
    d = _get_provedor("brasilapi", f"https://brasilapi.com.br/api/cep/v2/{cep}")
    if isinstance(d, dict) and d.get('cep'):
        return {
            "encontrado": True,
            "uf": (d.get('state') or '').upper().strip(),
            "localidade": normalizar_municipio(d.get('city')),
            "logradouro": (d.get('street') or '').upper().strip(),
            "bairro": (d.get('neighborhood') or '').upper().strip(),
        }
    return None

def _viacep_get(path, timeout=5):
    """GET genérico no ViaCEP (usado pela busca por logradouro), com descanso/ritmo."""
    return _get_provedor("viacep", f"https://viacep.com.br/ws/{path}", timeout=timeout)

def validar_cep_dne(cep):
    """Consulta o CEP nas APIs públicas, REVEZANDO ViaCEP e BrasilAPI para dividir
    a carga, com ritmo controlado e descanso por provedor. Cache por CEP.

    Retorna {encontrado, uf, localidade, logradouro, bairro}.
    DESACOPLADA: trocar pela API CEP oficial dos Correios mantém o formato.
    """
    vazio = {"encontrado": False, "uf": "", "localidade": "", "logradouro": "", "bairro": ""}
    cep_limpo = limpar_cep(cep)
    if len(cep_limpo) != 8:
        return vazio
    if cep_limpo in _cache_dne:
        return _cache_dne[cep_limpo]

    if _rr["i"] % 2 == 0:
        ordem = (_consultar_viacep, _consultar_brasilapi)
    else:
        ordem = (_consultar_brasilapi, _consultar_viacep)
    _rr["i"] += 1

    resultado = ordem[0](cep_limpo) or ordem[1](cep_limpo) or dict(vazio)
    # Só guarda no cache se confiável: achou, OU os dois provedores estão saudáveis
    # (evita "fixar" como não-encontrado um CEP que só falhou por bloqueio temporário).
    confiavel = resultado["encontrado"] or (
        _provedor_disponivel("viacep") and _provedor_disponivel("brasilapi"))
    if confiavel:
        _cache_dne[cep_limpo] = resultado
    return resultado

def algum_provedor_em_descanso():
    """True se algum provedor está em cooldown (para avisar na tela)."""
    return not (_provedor_disponivel("viacep") and _provedor_disponivel("brasilapi"))

def prevalidar_ceps(ceps, progresso=None, max_workers=5):
    """Pré-aquece o cache validando CEPs ÚNICOS (dedupe) com concorrência BAIXA e
    ritmo controlado — gentil com as APIs para não tomar bloqueio.
    `progresso(feitos, total)` é opcional.
    """
    unicos = sorted({
        limpar_cep(c) for c in ceps
        if len(limpar_cep(c)) == 8 and limpar_cep(c) not in _cache_dne
    })
    total = len(unicos)
    if progresso:
        progresso(0, total)
    if not total:
        return
    feitos = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futuros = [ex.submit(validar_cep_dne, c) for c in unicos]
        for _ in concurrent.futures.as_completed(futuros):
            feitos += 1
            if progresso and (feitos % 10 == 0 or feitos == total):
                progresso(feitos, total)

def buscar_cep_fallback(endereco, complemento, bairro, municipio, uf):
    """Tenta buscar CEP usando várias combinações de endereço - SÓ retorna se encontrar no mesmo UF"""
    uf_upper = uf.upper().strip() if uf else ''
    municipio_upper = municipio.upper().strip() if municipio else ''
    endereco_upper = endereco.upper() if endereco else ''
    partes = endereco_upper.split() if endereco else []
    
    buscas = []
    
    # 1. Bairro + cidade + UF
    buscas.append(f"{bairro} {municipio} {uf}".strip())
    
    # 2. Palavras do endereço + cidade + UF
    if len(partes) >= 2:
        for i in range(1, min(4, len(partes))):
            palavras = ' '.join(partes[:i])
            if len(palavras) > 5:
                buscas.append(f"{palavras} {municipio} {uf}".strip())
    
    # 3. Endereço completo
    buscas.append(f"{endereco} {bairro} {municipio} {uf}".strip())
    
    # 4. Padrões específicos do DF (SHIN, SHIS, SHIGS, SMPW, QI, QN, etc)
    padroes_df = ['SHIN', 'SHIS', 'SHIGS', 'SMPW', 'SQN', 'SQS', 'SRES', 'SQSW', 'SQNW', 'QI ', 'QN ', 'QS ']
    for padrao in padroes_df:
        if padrao in endereco_upper:
            idx = endereco_upper.find(padrao)
            parte = endereco[idx:idx+10].strip()
            if parte:
                buscas.append(f"{parte} {municipio} {uf}".strip())
                buscas.append(f"{parte} DF".strip())
                # Também tentar só a área (SHIN, SHIS, etc)
                if 'SHIN' in parte.upper():
                    buscas.append("SHIN DF")
                    buscas.append("ASA NORTE DF")
                elif 'SHIS' in parte.upper():
                    buscas.append("SHIS DF")
                    buscas.append("ASA SUL DF")
                elif 'QI ' in parte.upper() or 'QN ' in parte.upper():
                    buscas.append("ASA NORTE DF")
                    buscas.append("ASA SUL DF")
    
    # 5. Sem "QUADRA" ou "CONDOMINIO" + bairro + cidade + UF
    for palavra in ['QUADRA', 'CONDOMINIO', 'EDIFICIO', 'VILA', 'SETOR']:
        if palavra in endereco_upper:
            temp = endereco_upper.replace(palavra, '').strip()
            if len(temp) > 5:
                buscas.append(f"{temp} {bairro} {municipio} {uf}".strip())
    
    # 6. só a primeira palavra significativa do endereço + DF
    if partes:
        buscas.append(f"{partes[0]} DF".strip())
    
    # 7. Cidade + UF
    buscas.append(f"{municipio} {uf}".strip())
    
    for query in buscas:
        if not query or len(query) < 4:
            continue
        try:
            url = f"https://ceprua.com.br/api/buscar?q={query}"
            res = requests.get(url, timeout=3)
            if res.status_code == 200:
                dados = res.json()
                if 'resultados' in dados and dados['resultados']:
                    for item in dados['resultados']:
                        cep = item.get('cep', '').replace('-', '')
                        item_uf = item.get('uf', '').upper().strip()
                        if len(cep) == 8 and item_uf == uf_upper:
                            return cep
                if 'redirect' in dados:
                    cep = dados['redirect'].get('cep', '')
                    if cep:
                        return cep.replace('-', '')
        except:
            pass
    
    return ''

ABREVIACOES_END = {
    r'\bR\b': 'RUA', r'\bAV\b': 'AVENIDA', r'\bQD\b': 'QUADRA', r'\bQN\b': 'QUADRA',
    r'\bCS\b': 'CASA', r'\bAP\b': 'APARTAMENTO', r'\bAPTO\b': 'APARTAMENTO',
    r'\bBL\b': 'BLOCO', r'\bLT\b': 'LOTE', r'\bCONJ\b': 'CONJUNTO', r'\bCJ\b': 'CONJUNTO',
}
_GENERICOS = {'QUADRA', 'RUA', 'AVENIDA', 'SETOR', 'CONJUNTO', 'BLOCO', 'CASA',
              'LOTE', 'APARTAMENTO', 'DE', 'DO', 'DA', 'DAS', 'DOS', 'E'}

def _sem_acento(s):
    tab = str.maketrans('ÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ', 'AAAAAEEEEIIIIOOOOOUUUUC')
    return str(s or '').translate(tab)

def normalizar_endereco(texto):
    """Uppercase, sem acento, expande abreviações comuns e remove palavras duplicadas."""
    t = _sem_acento(str(texto or '').upper().strip())
    for padrao, sub in ABREVIACOES_END.items():
        t = re.sub(padrao, sub, t)
    out = []
    for p in t.split():
        if not out or p != out[-1]:
            out.append(p)
    return re.sub(r'\s+', ' ', ' '.join(out)).strip()

def _tokens_significativos(texto):
    """Tokens úteis para casar logradouros (ignora palavras genéricas)."""
    return {t for t in normalizar_endereco(texto).split() if t not in _GENERICOS}

# Cache da busca por logradouro no ViaCEP
_cache_logr = {}

def buscar_cep_por_logradouro(uf, cidade, logradouro):
    """Busca endereço->CEP na base oficial via ViaCEP (/ws/UF/Cidade/Logradouro/json/).
    Retorna lista de dicts {cep, logradouro, bairro} ou []. Desacoplada (trocar pela
    API CEP oficial dos Correios no futuro, se necessário).
    """
    uf = (uf or '').upper().strip()
    cidade = normalizar_municipio(cidade)
    logradouro = (logradouro or '').strip()
    if len(uf) != 2 or len(cidade) < 3 or len(logradouro) < 3:
        return []
    chave = (uf, cidade, logradouro.upper())
    if chave in _cache_logr:
        return _cache_logr[chave]
    res = []
    caminho = f"{uf}/{urllib.parse.quote(cidade)}/{urllib.parse.quote(logradouro)}/json/"
    d = _viacep_get(caminho, timeout=5)
    if isinstance(d, list):
        for x in d:
            cep = limpar_cep(x.get('cep', ''))
            if len(cep) == 8:
                res.append({
                    "cep": cep,
                    "logradouro": (x.get('logradouro') or '').upper(),
                    "bairro": (x.get('bairro') or '').upper(),
                })
    _cache_logr[chave] = res
    return res

def _termos_busca_logradouro(endereco):
    """Termos de busca para o ViaCEP, do mais específico ao mais amplo (>= 3 chars)."""
    n = normalizar_endereco(endereco)
    termos = []
    if n:
        termos.append(n)
        partes = n.split()
        if len(partes) >= 2 and partes[0] in {'QUADRA', 'RUA', 'AVENIDA', 'SETOR', 'CONJUNTO'}:
            termos.append(' '.join(partes[1:]))
        sig = [p for p in partes if p not in _GENERICOS]
        if sig:
            termos.append(' '.join(sig[:3]))
            termos.append(sig[0])
    vistos = []
    for t in termos:
        t = t.strip()
        if len(t) >= 3 and t not in vistos:
            vistos.append(t)
    return vistos

def _escolher_cep(resultados, bairro, endereco, complemento):
    """Escolhe o melhor CEP entre os resultados do ViaCEP (bairro + tokens bloco/número)."""
    if not resultados:
        return ''
    if len(resultados) == 1:
        return resultados[0]["cep"]
    bairro_n = normalizar_endereco(bairro)
    alvo = _tokens_significativos(f"{endereco} {complemento}")

    def score(r):
        s = 0
        rb = normalizar_endereco(r["bairro"])
        if bairro_n and rb and (bairro_n in rb or rb in bairro_n):
            s += 3
        s += 2 * len(alvo & _tokens_significativos(r["logradouro"]))
        return s

    ordenados = sorted(resultados, key=score, reverse=True)
    melhor = ordenados[0]
    if score(melhor) > 0 and score(melhor) > score(ordenados[1]):
        return melhor["cep"]
    # empate, mas todos na mesma quadra (5 primeiros dígitos) -> baixo risco
    if len({r["cep"][:5] for r in resultados}) == 1:
        return resultados[0]["cep"]
    return ''  # ambíguo -> deixa para fallback/revisão

def corrigir_cep_por_endereco(endereco, complemento, bairro, municipio, uf):
    """Acha o CEP correto pelo endereço, em cascata, sempre confirmando UF na base
    oficial. Ordem: ViaCEP busca-por-logradouro (CEP exato) -> CepRua (texto).
    Retorna CEP (8 díg) ou ''.
    """
    uf_alvo = (uf or '').upper().strip()

    # 1) ViaCEP busca por logradouro (endereço -> CEP exato, base DNE; UF garantida)
    for termo in _termos_busca_logradouro(endereco):
        cep = limpar_cep(_escolher_cep(
            buscar_cep_por_logradouro(uf_alvo, municipio, termo), bairro, endereco, complemento))
        if len(cep) == 8:
            return cep

    # 2) Fallback: CepRua (busca textual) confirmado no DNE com UF
    candidato = limpar_cep(buscar_cep_fallback(endereco, complemento, bairro, municipio, uf))
    if len(candidato) == 8:
        dne = validar_cep_dne(candidato)
        if dne["encontrado"] and (not uf_alvo or dne["uf"] == uf_alvo):
            return candidato
    return ''

def avaliar_registro(row):
    """Valida CPF e CEP de um registro contra a base oficial e tenta corrigir o
    CEP pelo endereço. Retorna um dict com cep_final, status e motivos da revisão.

    Status possíveis: "Válido", "Corrigido", "Revisão".
    """
    nome = str(row.get('NOME', '') or '')
    cpf_norm = normalizar_cpf(row.get('CPF', ''))
    endereco = str(row.get('ENDERECO', '') or '')
    bairro = str(row.get('BAIRRO', '') or '')
    municipio = str(row.get('MUNICIPIO', '') or '')
    uf = str(row.get('UF', '') or '')
    complemento = str(row.get('COMPLEMENTO DO ENDERECO', '') or '')

    cep_original = normalizar_cep(row.get('CEP', ''))
    uf_alvo = uf.upper().strip()

    motivos = []
    cpf_ok = cpf_valido(cpf_norm)
    if not cpf_ok:
        motivos.append(f"CPF/CNPJ com {len(limpar_digitos(cpf_norm))} dígitos (esperado 11 ou 14)")

    cep_final = cep_original
    cep_ok = False
    cep_status = "Revisão"

    if not cep_tem_formato_valido(cep_original):
        # >8 dígitos (ou impossível) -> tentar corrigir pelo endereço
        novo = corrigir_cep_por_endereco(endereco, complemento, bairro, municipio, uf)
        if novo:
            cep_final, cep_ok, cep_status = novo, True, "Corrigido"
        else:
            motivos.append("CEP com formato inválido (≠ 8 dígitos) e sem correção pelo endereço")
    else:
        dne = validar_cep_dne(cep_original)
        if dne["encontrado"] and (not uf_alvo or dne["uf"] == uf_alvo):
            cep_final, cep_ok, cep_status = cep_original, True, "Válido"
        else:
            novo = corrigir_cep_por_endereco(endereco, complemento, bairro, municipio, uf)
            if novo:
                cep_final, cep_ok, cep_status = novo, True, "Corrigido"
            elif not dne["encontrado"]:
                motivos.append("CEP não encontrado na base oficial")
            else:
                motivos.append(f"CEP pertence a {dne['uf']}, diferente da UF informada ({uf_alvo})")

    enviavel = cpf_ok and cep_ok
    if not enviavel:
        status = "Revisão"
    else:
        status = cep_status  # "Válido" ou "Corrigido"

    return {
        "nome": nome,
        "cpf": cpf_norm,
        "cpf_ok": cpf_ok,
        "cep_original": cep_original,
        "cep_final": cep_final,
        "endereco": endereco,
        "bairro": bairro,
        "municipio": municipio,
        "uf": uf,
        "complemento": complemento,
        "status": status,
        "enviavel": enviavel,
        "motivos": "; ".join(motivos),
    }

def linha_envio(r):
    """Linha no formato Correios (SR(A), 15 campos, sem aspas) usando o CEP final."""
    return (
        f'SR(A);{r["nome"].upper()};;;;{r["cpf"]};{r["cep_final"]};;'
        f'{r["endereco"].upper()};{r["bairro"].upper()};{r["municipio"].upper()};'
        f'{r["uf"].upper()};N;{r["complemento"].upper()[:40]};0'
    )


# ============================================
# INTERFACE
# ============================================

st.markdown("---")
st.markdown("## 📋 Formato da Planilha")
st.info("Colunas: CPF | MATRICULA | NOME | DATA NASCIMENTO | ENDERECO | COMPLEMENTO DO ENDERECO | BAIRRO | MUNICIPIO | UF | CEP")

# Download modelo
col1, col2 = st.columns([3, 1])
with col1:
    st.markdown("### 📤 Upload da Planilha")
with col2:
    modelo_df = pd.DataFrame(columns=CABEÇALHO_ESPERADO)
    modelo_buffer = excel_to_bytes_text(modelo_df)
    st.download_button("📥 Baixar Modelo", modelo_buffer.getvalue(), "modelo_planilha.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

uploaded_file = st.file_uploader("Carregar arquivo Excel (.xlsx)", type=['xlsx', 'xls'])

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file, dtype=str)
        df.columns = df.columns.str.strip()
        for col in df.columns:
            df[col] = df[col].astype(str)
        
        colunas_faltando = [c for c in CABEÇALHO_ESPERADO if c not in df.columns]
        if colunas_faltando:
            st.warning(f"Colunas faltando: {colunas_faltando}")
        else:
            st.success(f"✅ {len(df)} registros carregados")
        
        st.markdown("---")
        
        if st.button("✅ Processar Planilha", type="primary", use_container_width=True):
            progress_bar = st.progress(0)
            status_text = st.empty()
            aviso = st.empty()

            total = len(df)
            st.session_state.resultados = []
            status_count = {"Válido": 0, "Corrigido": 0, "Revisão": 0}

            # ---- Etapa 1/2: validar CEPs UNICOS (dedupe + ritmo controlado) ----
            ceps_col = df['CEP'] if 'CEP' in df.columns else []
            def _prog(feitos, tot):
                progress_bar.progress((feitos / tot) if tot else 1.0)
                status_text.text(f"Etapa 1/2 - Validando CEPs ({feitos}/{tot})")
                if algum_provedor_em_descanso():
                    aviso.warning("Uma das APIs de CEP esta limitando o ritmo agora; "
                                  "alternando para a outra e seguindo mais devagar.")
                else:
                    aviso.empty()
            status_text.text("Etapa 1/2 - Validando CEPs...")
            prevalidar_ceps((normalizar_cep(c) for c in ceps_col), progresso=_prog)

            # ---- Etapa 2/2: montar resultados ----
            progress_bar.progress(0)
            for idx, row in df.iterrows():
                r = avaliar_registro(row)
                status_count[r["status"]] = status_count.get(r["status"], 0) + 1
                st.session_state.resultados.append(r)
                if (idx + 1) % 25 == 0 or (idx + 1) == total:
                    progress_bar.progress((idx + 1) / total)
                    status_text.text(f"Etapa 2/2 - Gerando arquivos ({idx + 1}/{total})")

            aviso.empty()
            status_text.text("✅ Concluído!")

            resultados = st.session_state.resultados
            bloco_size = 999

            # ===== Arquivo de ENVIO (validos + corrigidos, com CEP final) =====
            enviaveis = [r for r in resultados if r["enviavel"]]
            linhas_envio = [linha_envio(r) for r in enviaveis]

            st.session_state.blocos_envio = []
            for i in range(0, len(linhas_envio), bloco_size):
                bloco = '\ufeff' + '\n'.join(linhas_envio[i:i+bloco_size])
                st.session_state.blocos_envio.append(bloco)

            st.session_state.csv_envio = '\ufeff' + '\n'.join(linhas_envio)
            st.session_state.count_envio = len(linhas_envio)

            # ===== Arquivo de REVISAO MANUAL (xlsx com motivos) =====
            colunas_revisao = ["NOME", "CPF", "CEP_INFORMADO", "ENDERECO", "BAIRRO",
                               "MUNICIPIO", "UF", "COMPLEMENTO", "MOTIVO"]
            revisao = [r for r in resultados if r["status"] == "Revisão"]
            df_revisao = pd.DataFrame([{
                "NOME": r["nome"], "CPF": r["cpf"], "CEP_INFORMADO": r["cep_original"],
                "ENDERECO": r["endereco"], "BAIRRO": r["bairro"], "MUNICIPIO": r["municipio"],
                "UF": r["uf"], "COMPLEMENTO": r["complemento"], "MOTIVO": r["motivos"],
            } for r in revisao], columns=colunas_revisao)
            for col in df_revisao.columns:
                df_revisao[col] = df_revisao[col].astype(str)
            st.session_state.buffer_revisao = excel_to_bytes_text(df_revisao).getvalue()
            st.session_state.count_revisao = len(df_revisao)

            # ===== Base CONSOLIDADA (todos os registros + resultado) =====
            df_consolidada = pd.DataFrame([{
                "NOME": r["nome"], "CPF": r["cpf"], "CEP_INFORMADO": r["cep_original"],
                "CEP_FINAL": r["cep_final"], "STATUS": r["status"], "ENDERECO": r["endereco"],
                "BAIRRO": r["bairro"], "MUNICIPIO": r["municipio"], "UF": r["uf"],
                "COMPLEMENTO": r["complemento"], "MOTIVO": r["motivos"],
            } for r in resultados])
            for col in df_consolidada.columns:
                df_consolidada[col] = df_consolidada[col].astype(str)
            st.session_state.buffer_consolidada = excel_to_bytes_text(df_consolidada).getvalue()

            st.session_state.status_count = status_count
            st.session_state.processado = True
    
    except Exception as e:
        st.error(f"Erro: {e}")

# ========================================
# MOSTRAR RESULTADOS (se já processado)
# ========================================
if st.session_state.get('processado', False):
    sc = st.session_state.get('status_count', {})
    st.markdown("---")
    st.markdown("## Resultados")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total", len(st.session_state.resultados))
    c2.metric("Validos", sc.get("Válido", 0))
    c3.metric("Corrigidos", sc.get("Corrigido", 0))
    c4.metric("Revisao manual", sc.get("Revisão", 0))

    st.markdown("---")
    st.markdown("## Downloads")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.download_button(
            f"Envio Correios ({st.session_state.count_envio})",
            st.session_state.csv_envio, "envio_correios.csv",
            "text/csv; charset=utf-8-sig", use_container_width=True)
    with col2:
        st.download_button(
            f"Revisao manual ({st.session_state.count_revisao})",
            st.session_state.buffer_revisao, "revisao_manual.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with col3:
        st.download_button(
            f"Base consolidada ({len(st.session_state.resultados)})",
            st.session_state.buffer_consolidada, "base_consolidada.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

    if st.session_state.count_revisao:
        st.warning(
            f"{st.session_state.count_revisao} registro(s) precisam de revisao manual "
            "(CPF invalido, CEP nao encontrado ou UF divergente). Veja o arquivo de revisao."
        )

    st.markdown("### Blocos de 999 (Envio Correios)")
    for i, bloco in enumerate(st.session_state.get('blocos_envio', [])):
        ini = i * 999 + 1
        fim = min(i * 999 + 999, st.session_state.count_envio)
        st.download_button(
            f"Bloco {i+1} (linhas {ini}-{fim})", bloco,
            f"envio_correios_bloco_{i+1}.csv",
            "text/csv; charset=utf-8-sig", use_container_width=True)

    st.success("Processamento concluido!")

st.markdown("---")
st.caption("CEP Paperclipe v6.0 - Validacao DNE (ViaCEP) + revisao manual")
